"""
Daily Route Planner - Main Application Module

A comprehensive route planning web application with:
- Route optimization using Nearest Neighbor algorithm
- Points of Interest discovery
- Multi-user system (Individual, Business, Driver)
- Real-time GPS tracking via WebSocket
- AI predictions using linear regression
- Export capabilities (PDF, Excel, CSV, QR)
- PWA support for mobile installation

Author: Chioma Okoye
Year: 2026
"""

import os
import json
import math
import time
import sqlite3
import hashlib
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from flask import Flask, request, jsonify, render_template, send_file, make_response, g, session
from flask_sock import Sock
import numpy as np
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import qrcode

# ============================================================================
# FLASK APPLICATION INITIALIZATION
# ============================================================================

app = Flask(__name__)
app.config['SECRET_KEY'] = 'daily-route-planner-2026-chioma-okoye'
app.config['DATABASE'] = os.path.join(os.path.dirname(__file__), 'route_planner.db')

# Initialize Flask-Sock for WebSocket real-time communication
sock = Sock(app)

# Global storage for driver locations (in production, use Redis)
driver_locations = {}

# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================

def get_db():
    """
    Get database connection for current request context.
    
    WHY: Using g object ensures one connection per request (Flask best practice).
    COMPLEXITY: O(1) - Simple attribute access.
    
    Returns:
        sqlite3.Connection: Database connection object
    
    Example:
        >>> db = get_db()
        >>> cursor = db.cursor()
    """
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception):
    """
    Close database connection at end of request.
    
    WHY: Prevents connection leaks and resource exhaustion.
    COMPLEXITY: O(1) - Single close operation.
    
    Args:
        exception: Any exception that occurred during request
    """
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    """
    Initialize database with all required tables and demo data.
    
    WHY: Centralized schema management for reproducibility.
    COMPLEXITY: O(1) - Fixed number of CREATE TABLE statements.
    
    Tables created:
        - users: Multi-user system with role-based access
        - saved_locations: User's frequently used locations
        - routes: Saved route information
        - route_stops: Individual stops within routes
        - points_of_interest: POIs along routes
        - driver_performance: Driver ratings and metrics
        - trip_reports: Historical trip data for ML training
    
    Example:
        >>> init_db()
        # Creates all tables and demo accounts
    """
    db = sqlite3.connect(app.config['DATABASE'])
    cursor = db.cursor()
    
    # Users table - supports three user types
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            email TEXT,
            user_type TEXT NOT NULL CHECK(user_type IN ('individual', 'business', 'driver')),
            business_name TEXT,
            vehicle_type TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Saved locations for quick access
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS saved_locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            category TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # Routes table - stores optimized routes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS routes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            total_distance REAL,
            total_duration REAL,
            fuel_consumption REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # Route stops - individual waypoints
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS route_stops (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            route_id INTEGER NOT NULL,
            stop_order INTEGER NOT NULL,
            name TEXT NOT NULL,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            arrival_time TEXT,
            departure_time TEXT,
            FOREIGN KEY (route_id) REFERENCES routes(id)
        )
    ''')
    
    # Points of Interest along routes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS points_of_interest (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            route_id INTEGER,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            rating REAL,
            distance_from_route REAL,
            FOREIGN KEY (route_id) REFERENCES routes(id)
        )
    ''')
    
    # Driver performance metrics
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS driver_performance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            driver_id INTEGER NOT NULL,
            business_id INTEGER NOT NULL,
            total_trips INTEGER DEFAULT 0,
            total_distance REAL DEFAULT 0,
            average_rating REAL DEFAULT 5.0,
            on_time_percentage REAL DEFAULT 100.0,
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (driver_id) REFERENCES users(id),
            FOREIGN KEY (business_id) REFERENCES users(id)
        )
    ''')
    
    # Trip reports for ML training and analytics
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trip_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            driver_id INTEGER,
            route_id INTEGER,
            distance REAL,
            duration REAL,
            num_stops INTEGER,
            departure_hour INTEGER,
            day_of_week INTEGER,
            month INTEGER,
            is_weekend BOOLEAN,
            is_rush_hour BOOLEAN,
            fuel_consumed REAL,
            rating REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (driver_id) REFERENCES users(id),
            FOREIGN KEY (route_id) REFERENCES routes(id)
        )
    ''')
    
    # Create demo accounts
    demo_users = [
        ('business_owner', hash_password('pass123'), 'business@example.com', 'business', 'Logistics Co.', None),
        ('driver_sarah', hash_password('pass123'), 'sarah@example.com', 'driver', None, 'Van'),
        ('traveler_john', hash_password('pass123'), 'john@example.com', 'individual', None, None),
    ]
    
    for user in demo_users:
        try:
            cursor.execute('''
                INSERT INTO users (username, password_hash, email, user_type, business_name, vehicle_type)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', user)
        except sqlite3.IntegrityError:
            pass  # User already exists
    
    # Add sample trip data for ML training
    add_sample_trip_data(cursor)
    
    db.commit()
    db.close()


def hash_password(password):
    """
    Hash password using SHA-256 for secure storage.
    
    WHY: SHA-256 provides adequate security for student project.
         In production, use bcrypt or argon2.
    COMPLEXITY: O(n) where n is password length.
    
    Args:
        password (str): Plain text password
    
    Returns:
        str: Hexadecimal hash string
    
    Example:
        >>> hash_password('pass123')
        'ef92b778bafe771e89245b89ecbc08a44a4e166c06659911881f383d4473e94f'
    """
    return hashlib.sha256(password.encode()).hexdigest()


def add_sample_trip_data(cursor):
    """
    Add historical trip data for ML model training.
    
    WHY: Linear regression needs training data to make predictions.
    COMPLEXITY: O(n) where n is number of sample records.
    
    Args:
        cursor: SQLite cursor object
    
    Note: Data simulates realistic traffic patterns:
          - Rush hours (7-9 AM, 5-7 PM) have longer durations
          - Weekends have lighter traffic
          - More stops increase total time
    """
    sample_trips = [
        # (distance, duration, stops, hour, dow, month, weekend, rush, fuel, rating)
        (50.0, 65.0, 3, 8, 1, 3, 0, 1, 5.5, 4.5),   # Monday rush
        (50.0, 55.0, 3, 11, 1, 3, 0, 0, 5.0, 4.8),  # Monday off-peak
        (75.0, 95.0, 5, 17, 2, 3, 0, 1, 8.0, 4.2),  # Tuesday evening rush
        (30.0, 35.0, 2, 10, 3, 3, 0, 0, 3.2, 4.9),  # Wednesday morning
        (100.0, 125.0, 7, 14, 4, 3, 0, 0, 10.5, 4.6), # Thursday afternoon
        (60.0, 80.0, 4, 18, 5, 3, 0, 1, 6.5, 4.3),  # Friday rush
        (45.0, 50.0, 3, 10, 6, 3, 1, 0, 4.8, 4.7),  # Saturday
        (40.0, 45.0, 2, 14, 0, 3, 1, 0, 4.2, 4.8),  # Sunday
        (55.0, 70.0, 4, 7, 1, 4, 0, 1, 6.0, 4.4),   # Monday morning rush
        (80.0, 100.0, 6, 16, 2, 4, 0, 1, 8.5, 4.1), # Tuesday late afternoon
        (35.0, 40.0, 2, 12, 3, 4, 0, 0, 3.8, 4.9),  # Wednesday noon
        (90.0, 110.0, 6, 9, 4, 4, 0, 1, 9.5, 4.5),  # Thursday morning
        (65.0, 75.0, 4, 19, 5, 4, 0, 1, 7.0, 4.4),  # Friday evening
        (50.0, 55.0, 3, 11, 6, 4, 1, 0, 5.2, 4.8),  # Saturday late morning
        (42.0, 48.0, 3, 15, 0, 4, 1, 0, 4.5, 4.7),  # Sunday afternoon
        (70.0, 90.0, 5, 8, 1, 5, 0, 1, 7.5, 4.3),   # Monday rush
        (55.0, 60.0, 3, 13, 2, 5, 0, 0, 5.8, 4.6),  # Tuesday afternoon
        (48.0, 52.0, 3, 10, 3, 5, 0, 0, 5.0, 4.8),  # Wednesday
        (85.0, 105.0, 6, 17, 4, 5, 0, 1, 9.0, 4.2), # Thursday rush
        (60.0, 68.0, 4, 11, 5, 5, 0, 0, 6.2, 4.5),  # Friday late morning
    ]
    
    for trip in sample_trips:
        cursor.execute('''
            INSERT INTO trip_reports 
            (distance, duration, num_stops, departure_hour, day_of_week, month, 
             is_weekend, is_rush_hour, fuel_consumed, rating)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', trip)


# ============================================================================
# ROUTE OPTIMIZATION ALGORITHMS
# ============================================================================

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Calculate great-circle distance between two GPS coordinates.
    
    WHY: Haversine formula accounts for Earth's curvature, providing accurate
         distances for route planning. Alternative: Vincenty formula (more accurate
         but computationally expensive).
    COMPLEXITY: O(1) - Fixed number of mathematical operations.
    
    Args:
        lat1 (float): Latitude of point 1 in degrees
        lon1 (float): Longitude of point 1 in degrees
        lat2 (float): Latitude of point 2 in degrees
        lon2 (float): Longitude of point 2 in degrees
    
    Returns:
        float: Distance in kilometers
    
    Example:
        >>> haversine_distance(51.5074, -0.1278, 48.8566, 2.3522)
        343.7  # London to Paris approximately
    """
    # Earth's radius in kilometers (mean radius)
    R = 6371.0
    
    # Convert degrees to radians for trigonometric functions
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    # Haversine formula: a = sin²(Δφ/2) + cos φ1 ⋅ cos φ2 ⋅ sin²(Δλ/2)
    a = math.sin(delta_lat / 2) ** 2 + \
        math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon / 2) ** 2
    
    # c = 2 ⋅ atan2(√a, √(1−a))
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    # Distance = radius × central angle
    return R * c


def nearest_neighbor_optimization(stops):
    """
    Optimize route using Nearest Neighbor heuristic algorithm.
    
    WHY: Nearest Neighbor provides good solutions quickly (O(n²)) compared to
         brute force O(n!). For 10+ stops, this is practical. Alternative approaches:
         - Genetic Algorithm: Better quality but slower
         - Simulated Annealing: Good balance but complex
         - 2-opt improvement: Can refine NN results
    COMPLEXITY: O(n²) where n is number of stops.
    
    Args:
        stops (list): List of dicts with 'name', 'latitude', 'longitude'
    
    Returns:
        dict: Optimized route with order, distances, and statistics
    
    Example:
        >>> stops = [
        ...     {'name': 'A', 'latitude': 51.5, 'longitude': -0.1},
        ...     {'name': 'B', 'latitude': 51.6, 'longitude': -0.2},
        ...     {'name': 'C', 'latitude': 51.4, 'longitude': -0.3}
        ... ]
        >>> result = nearest_neighbor_optimization(stops)
        >>> 'optimized_order' in result
        True
    """
    if len(stops) <= 1:
        return {
            'optimized_order': stops,
            'total_distance': 0,
            'total_duration': 0,
            'distance_saved': 0,
            'improvement_percent': 0
        }
    
    # Start from first stop (could be improved by trying multiple starting points)
    unvisited = stops[1:].copy()  # All stops except first
    optimized = [stops[0]]  # Start with first stop
    
    # Greedy approach: always visit nearest unvisited stop
    while unvisited:
        current = optimized[-1]
        
        # Find nearest neighbor - O(n) operation
        nearest = None
        min_distance = float('inf')
        
        for stop in unvisited:
            dist = haversine_distance(
                current['latitude'], current['longitude'],
                stop['latitude'], stop['longitude']
            )
            if dist < min_distance:
                min_distance = dist
                nearest = stop
        
        optimized.append(nearest)
        unvisited.remove(nearest)
    
    # Calculate total distance of optimized route
    total_distance = 0
    segments = []
    
    for i in range(len(optimized) - 1):
        seg_dist = haversine_distance(
            optimized[i]['latitude'], optimized[i]['longitude'],
            optimized[i + 1]['latitude'], optimized[i + 1]['longitude']
        )
        total_distance += seg_dist
        segments.append({
            'from': optimized[i]['name'],
            'to': optimized[i + 1]['name'],
            'distance': round(seg_dist, 2)
        })
    
    # Calculate original distance (order as input) for comparison
    original_distance = 0
    for i in range(len(stops) - 1):
        original_distance += haversine_distance(
            stops[i]['latitude'], stops[i]['longitude'],
            stops[i + 1]['latitude'], stops[i + 1]['longitude']
        )
    
    # Duration estimation: 50 km/h average + 10% buffer + 15 min per stop
    avg_speed = 50  # km/h
    buffer = 1.1  # 10% buffer for traffic, lights, etc.
    stop_time = 0.25  # 15 minutes per intermediate stop in hours
    
    driving_time = (total_distance / avg_speed) * buffer
    total_stop_time = stop_time * (len(optimized) - 2) if len(optimized) > 2 else 0
    total_duration = driving_time + total_stop_time
    
    # Convert to minutes for display
    total_duration_minutes = total_duration * 60
    
    distance_saved = original_distance - total_distance
    improvement = (distance_saved / original_distance * 100) if original_distance > 0 else 0
    
    return {
        'optimized_order': optimized,
        'total_distance': round(total_distance, 2),
        'total_duration': round(total_duration_minutes, 1),
        'segments': segments,
        'original_distance': round(original_distance, 2),
        'distance_saved': round(distance_saved, 2),
        'improvement_percent': round(improvement, 1)
    }


# ============================================================================
# AI PREDICTIONS USING LINEAR REGRESSION
# ============================================================================

class LinearRegressionModel:
    """
    Simple Linear Regression using Normal Equation.
    
    WHY: Normal Equation provides closed-form solution without iteration.
         Suitable for small datasets (< 1000 samples). For larger datasets,
         Gradient Descent would be more efficient.
    COMPLEXITY: O(n³) for matrix inversion, but n is small (number of features).
    
    Features used:
        - distance: Total route distance (km)
        - stops: Number of intermediate stops
        - hour: Departure hour (0-23)
        - day_of_week: Day index (0=Sunday, 6=Saturday)
        - month: Month of year (1-12)
        - is_weekend: Boolean (1 if Sat/Sun, 0 otherwise)
        - is_rush_hour: Boolean (1 if 7-9 AM or 5-7 PM, 0 otherwise)
    
    Example:
        >>> model = LinearRegressionModel()
        >>> model.train(X_train, y_train)
        >>> prediction = model.predict([50, 3, 8, 1, 3, 0, 1])
    """
    
    def __init__(self):
        """Initialize model with empty weights."""
        self.weights = None
        self.bias = 0
        self.is_trained = False
    
    def _add_bias(self, X):
        """
        Add bias column (ones) to feature matrix.
        
        WHY: Allows model to learn intercept term.
        COMPLEXITY: O(n*m) where n=samples, m=features.
        
        Args:
            X (np.ndarray): Feature matrix
            
        Returns:
            np.ndarray: Feature matrix with bias column
        """
        return np.column_stack([np.ones(len(X)), X])
    
    def train(self, X, y):
        """
        Train model using Normal Equation: θ = (X^T X)^(-1) X^T y
        
        WHY: Closed-form solution avoids hyperparameter tuning needed for
             Gradient Descent. Computationally feasible for small datasets.
        COMPLEXITY: O(m³) where m is number of features (matrix inversion).
        
        Args:
            X (np.ndarray): Training features, shape (n_samples, n_features)
            y (np.ndarray): Target values, shape (n_samples,)
        
        Example:
            >>> X = np.array([[50, 3, 8], [75, 5, 17]])
            >>> y = np.array([65, 95])
            >>> model.train(X, y)
        """
        # Add bias column
        X_bias = self._add_bias(X)
        
        # Normal Equation: θ = (X^T X)^(-1) X^T y
        # Step 1: Compute X^T X
        XtX = X_bias.T @ X_bias
        
        # Step 2: Compute inverse with regularization for numerical stability
        # Adding small value to diagonal prevents singular matrix
        regularization = 0.01 * np.eye(XtX.shape[0])
        XtX_inv = np.linalg.inv(XtX + regularization)
        
        # Step 3: Compute X^T y
        Xty = X_bias.T @ y
        
        # Step 4: Compute final weights
        theta = XtX_inv @ Xty
        
        self.bias = theta[0]
        self.weights = theta[1:]
        self.is_trained = True
    
    def predict(self, X):
        """
        Make predictions using trained model.
        
        WHY: Linear combination of features with learned weights.
        COMPLEXITY: O(m) where m is number of features.
        
        Args:
            X (np.ndarray or list): Features for prediction
            
        Returns:
            float: Predicted value (duration in minutes)
        
        Raises:
            ValueError: If model hasn't been trained
            
        Example:
            >>> model.predict([50, 3, 8, 1, 3, 0, 1])
            65.3
        """
        if not self.is_trained:
            raise ValueError("Model must be trained before prediction")
        
        X = np.array(X)
        if X.ndim == 1:
            X = X.reshape(1, -1)
        
        # Prediction: y = θ₀ + θ₁x₁ + θ₂x₂ + ... + θₘxₘ
        return X @ self.weights + self.bias
    
    def get_feature_importance(self):
        """
        Return feature importance based on weight magnitudes.
        
        WHY: Understanding which features most affect predictions helps
             with model interpretation and feature engineering.
        COMPLEXITY: O(m) where m is number of features.
        
        Returns:
            dict: Feature names mapped to importance scores
        
        Example:
            >>> model.get_feature_importance()
            {'distance': 0.8, 'stops': 0.3, ...}
        """
        features = ['distance', 'stops', 'hour', 'day_of_week', 
                   'month', 'is_weekend', 'is_rush_hour']
        
        if self.weights is None:
            return {f: 0 for f in features}
        
        # Normalize weights to show relative importance
        total = np.sum(np.abs(self.weights))
        if total == 0:
            return {f: 0 for f in features}
        
        return {features[i]: abs(w) / total for i, w in enumerate(self.weights)}


def get_traffic_model(db_cursor):
    """
    Load or train traffic prediction model from database.
    
    WHY: Model persists across requests, avoiding retraining.
    COMPLEXITY: O(n) to fetch data + O(m³) to train model.
    
    Args:
        db_cursor: SQLite database cursor
    
    Returns:
        LinearRegressionModel: Trained prediction model
    
    Example:
        >>> model = get_traffic_model(db.cursor())
        >>> prediction = model.predict([50, 3, 8, 1, 3, 0, 1])
    """
    # Fetch historical trip data
    db_cursor.execute('''
        SELECT distance, num_stops, departure_hour, day_of_week, month,
               is_weekend, is_rush_hour, duration
        FROM trip_reports
    ''')
    
    rows = db_cursor.fetchall()
    
    if len(rows) < 5:
        # Not enough data, return default model
        model = LinearRegressionModel()
        # Default weights based on domain knowledge
        model.weights = np.array([1.0, 3.0, 0.5, 0.1, 0.05, -2.0, 8.0])
        model.bias = 10.0
        model.is_trained = True
        return model
    
    # Prepare training data
    X = np.array([[row[0], row[1], row[2], row[3], row[4], row[5], row[6]] 
                 for row in rows])
    y = np.array([row[7] for row in rows])
    
    # Train model
    model = LinearRegressionModel()
    model.train(X, y)
    
    return model


def predict_travel_time(model, distance, num_stops, departure_hour=None):
    """
    Predict travel duration with optimal departure time suggestions.
    
    WHY: Helps users plan trips efficiently by avoiding traffic.
    COMPLEXITY: O(1) for single prediction, O(24) for hourly analysis.
    
    Args:
        model (LinearRegressionModel): Trained prediction model
        distance (float): Route distance in km
        num_stops (int): Number of intermediate stops
        departure_hour (int, optional): Specific hour for prediction
    
    Returns:
        dict: Predictions including optimal times and fuel estimates
    
    Example:
        >>> result = predict_travel_time(model, 50, 3, 8)
        >>> 'predicted_duration' in result
        True
    """
    current_hour = datetime.now().hour
    current_day = datetime.now().weekday()
    current_month = datetime.now().month
    is_weekend = 1 if current_day >= 5 else 0
    is_rush = 1 if current_hour in [7, 8, 17, 18] else 0
    
    # Base prediction for current conditions
    base_features = [distance, num_stops, current_hour, current_day, 
                    current_month, is_weekend, is_rush]
    base_prediction = model.predict(base_features)[0]
    
    # Analyze all 24 hours to find optimal departure times
    hourly_predictions = []
    for hour in range(24):
        is_rush_hour = 1 if hour in [7, 8, 17, 18] else 0
        features = [distance, num_stops, hour, current_day, 
                   current_month, is_weekend, is_rush_hour]
        pred = model.predict(features)[0]
        hourly_predictions.append({
            'hour': hour,
            'predicted_duration': round(pred, 1),
            'is_rush_hour': bool(is_rush_hour)
        })
    
    # Sort by duration and get top 3 best times
    sorted_hours = sorted(hourly_predictions, key=lambda x: x['predicted_duration'])
    best_times = sorted_hours[:3]
    
    # Format times for display
    for time_slot in best_times:
        hour = time_slot['hour']
        period = 'AM' if hour < 12 else 'PM'
        display_hour = hour if hour <= 12 else hour - 12
        if display_hour == 0:
            display_hour = 12
        time_slot['display_time'] = f"{display_hour}:00 {period}"
    
    # Fuel consumption prediction (liters)
    # Based on average consumption: 8L/100km for van, 6L/100km for car
    fuel_car = distance * 0.06
    fuel_van = distance * 0.08
    fuel_truck = distance * 0.12
    
    return {
        'predicted_duration': round(base_prediction, 1),
        'best_departure_times': best_times,
        'fuel_consumption': {
            'car': round(fuel_car, 2),
            'van': round(fuel_van, 2),
            'truck': round(fuel_truck, 2)
        },
        'current_conditions': {
            'hour': current_hour,
            'is_rush_hour': bool(is_rush),
            'is_weekend': bool(is_weekend)
        }
    }


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def generate_pdf_report(route_data, poi_data, predictions):
    """
    Generate PDF report using ReportLab.
    
    WHY: PDF provides universal format for sharing route plans.
    COMPLEXITY: O(n) where n is number of elements in report.
    
    Args:
        route_data (dict): Optimized route information
        poi_data (list): Points of interest along route
        predictions (dict): AI predictions
    
    Returns:
        bytes: PDF file content
    
    Example:
        >>> pdf_bytes = generate_pdf_report(route, pois, preds)
        >>> send_file(pdf_bytes, mimetype='application/pdf')
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title = Paragraph("Daily Route Planner - Trip Report", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Date
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    elements.append(Paragraph(f"Generated: {date_str}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Route Summary Table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Distance', f"{route_data['total_distance']} km"],
        ['Estimated Duration', f"{route_data['total_duration']} minutes"],
        ['Number of Stops', str(len(route_data['optimized_order']))],
        ['Distance Saved', f"{route_data['distance_saved']} km"],
        ['Improvement', f"{route_data['improvement_percent']}%"]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Route Stops
    elements.append(Paragraph("Route Stops (Optimized Order)", styles['Heading2']))
    stops_data = [['Order', 'Location', 'Latitude', 'Longitude']]
    
    for i, stop in enumerate(route_data['optimized_order']):
        stops_data.append([
            str(i + 1),
            stop['name'],
            str(stop['latitude']),
            str(stop['longitude'])
        ])
    
    stops_table = Table(stops_data)
    stops_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER')
    ]))
    elements.append(stops_table)
    elements.append(Spacer(1, 20))
    
    # AI Predictions
    elements.append(Paragraph("AI Travel Predictions", styles['Heading2']))
    pred_data = [
        ['Predicted Duration', f"{predictions['predicted_duration']} minutes"],
        ['Best Departure Time', predictions['best_departure_times'][0]['display_time']],
        ['Fuel (Car)', f"{predictions['fuel_consumption']['car']} L"],
        ['Fuel (Van)', f"{predictions['fuel_consumption']['van']} L"]
    ]
    
    pred_table = Table(pred_data)
    pred_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(pred_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return buffer.getvalue()


def generate_excel_report(route_data, poi_data, trip_history):
    """
    Generate Excel report using openpyxl.
    
    WHY: Excel allows further analysis and customization by users.
    COMPLEXITY: O(n) where n is number of rows written.
    
    Args:
        route_data (dict): Route information
        poi_data (list): Points of interest
        trip_history (list): Historical trips
    
    Returns:
        bytes: Excel file content
    
    Example:
        >>> excel_bytes = generate_excel_report(route, pois, history)
    """
    wb = Workbook()
    
    # Route Summary Sheet
    ws1 = wb.active
    ws1.title = "Route Summary"
    
    ws1['A1'] = "Daily Route Planner Report"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    headers = ['Metric', 'Value']
    ws1.append(headers)
    
    metrics = [
        ('Total Distance', f"{route_data['total_distance']} km"),
        ('Duration', f"{route_data['total_duration']} min"),
        ('Stops', len(route_data['optimized_order'])),
        ('Distance Saved', f"{route_data['distance_saved']} km"),
        ('Improvement', f"{route_data['improvement_percent']}%")
    ]
    
    for metric in metrics:
        ws1.append(metric)
    
    # Style header row
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Route Stops Sheet
    ws2 = wb.create_sheet("Route Stops")
    ws2.append(['Order', 'Name', 'Latitude', 'Longitude'])
    
    for i, stop in enumerate(route_data['optimized_order']):
        ws2.append([i + 1, stop['name'], stop['latitude'], stop['longitude']])
    
    # POI Sheet
    if poi_data:
        ws3 = wb.create_sheet("Points of Interest")
        ws3.append(['Category', 'Name', 'Distance from Route', 'Rating'])
        
        for poi in poi_data:
            ws3.append([
                poi.get('category', 'Unknown'),
                poi.get('name', 'Unknown'),
                poi.get('distance', 0),
                poi.get('rating', 0)
            ])
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def generate_qr_code(route_url):
    """
    Generate QR code for quick route access.
    
    WHY: QR codes enable easy sharing and mobile access.
    COMPLEXITY: O(n) where n is URL length.
    
    Args:
        route_url (str): URL to encode in QR code
    
    Returns:
        bytes: PNG image data
    
    Example:
        >>> qr_img = generate_qr_code('http://example.com/route/123')
    """
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4
    )
    qr.add_data(route_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    return buffer.getvalue()


# ============================================================================
# AUTHENTICATION DECORATORS
# ============================================================================

def login_required(f):
    """
    Decorator to require user authentication.
    
    WHY: Protects routes that need user context.
    COMPLEXITY: O(1) - Session lookup.
    
    Args:
        f: Flask view function
    
    Returns:
        Wrapped function with authentication check
    
    Example:
        >>> @app.route('/dashboard')
        ... @login_required
        ... def dashboard():
        ...     return render_template('dashboard.html')
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function


def role_required(allowed_roles):
    """
    Decorator to restrict access by user role.
    
    WHY: Role-based access control for business features.
    COMPLEXITY: O(1) - Role check.
    
    Args:
        allowed_roles (list): List of allowed user types
    
    Returns:
        Decorator function
    
    Example:
        >>> @app.route('/drivers')
        ... @role_required(['business'])
        ... def manage_drivers():
        ...     return render_template('drivers.html')
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_type' not in session:
                return jsonify({'error': 'Authentication required'}), 401
            
            if session['user_type'] not in allowed_roles:
                return jsonify({'error': 'Insufficient permissions'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ============================================================================
# FLASK ROUTES - AUTHENTICATION
# ============================================================================

@app.route('/')
def index():
    """
    Render main application page.
    
    WHY: Single-page application design for better UX.
    COMPLEXITY: O(1) - Template rendering.
    
    Returns:
        str: Rendered HTML template
    
    Example:
        >>> response = client.get('/')
        >>> response.status_code
        200
    """
    return render_template('index.html')


@app.route('/api/login', methods=['POST'])
def login():
    """
    Authenticate user and create session.
    
    WHY: Session-based auth is simple and effective for this scale.
    COMPLEXITY: O(1) - Single DB query.
    
    Request JSON:
        - username: User's username
        - password: User's password
    
    Returns:
        JSON: User info on success, error on failure
    
    Example:
        >>> response = client.post('/api/login', 
        ...     json={'username': 'traveler_john', 'password': 'pass123'})
        >>> response.json['success']
        True
    """
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    
    db = get_db()
    cursor = db.cursor()
    
    password_hash = hash_password(password)
    
    cursor.execute('''
        SELECT id, username, user_type, business_name, vehicle_type
        FROM users
        WHERE username = ? AND password_hash = ?
    ''', (username, password_hash))
    
    user = cursor.fetchone()
    
    if not user:
        return jsonify({'error': 'Invalid credentials'}), 401
    
    # Store user info in session
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['user_type'] = user['user_type']
    
    return jsonify({
        'success': True,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'user_type': user['user_type'],
            'business_name': user['business_name'],
            'vehicle_type': user['vehicle_type']
        }
    })


@app.route('/api/logout', methods=['POST'])
def logout():
    """
    Clear user session.
    
    WHY: Proper session cleanup for security.
    COMPLEXITY: O(1) - Session clear.
    
    Returns:
        JSON: Success confirmation
    
    Example:
        >>> response = client.post('/api/logout')
        >>> response.json['success']
        True
    """
    session.clear()
    return jsonify({'success': True})


@app.route('/api/register', methods=['POST'])
def register():
    """
    Register new user account.
    
    WHY: Allow users to create accounts for personalized experience.
    COMPLEXITY: O(1) - Single INSERT query.
    
    Request JSON:
        - username: Desired username
        - password: Password
        - email: Email address
        - user_type: 'individual', 'business', or 'driver'
        - business_name: (optional) For business accounts
        - vehicle_type: (optional) For driver accounts
    
    Returns:
        JSON: Success or error message
    
    Example:
        >>> response = client.post('/api/register',
        ...     json={'username': 'new_user', 'password': 'secure123',
        ...           'email': 'new@example.com', 'user_type': 'individual'})
    """
    data = request.get_json()
    
    required = ['username', 'password', 'email', 'user_type']
    if not all(k in data for k in required):
        return jsonify({'error': 'Missing required fields'}), 400
    
    if data['user_type'] not in ['individual', 'business', 'driver']:
        return jsonify({'error': 'Invalid user type'}), 400
    
    password_hash = hash_password(data['password'])
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute('''
            INSERT INTO users (username, password_hash, email, user_type, 
                             business_name, vehicle_type)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data['username'],
            password_hash,
            data['email'],
            data['user_type'],
            data.get('business_name'),
            data.get('vehicle_type')
        ))
        db.commit()
        
        return jsonify({'success': True, 'message': 'Account created'})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already exists'}), 400


# ============================================================================
# FLASK ROUTES - ROUTE OPTIMIZATION
# ============================================================================

@app.route('/api/optimize', methods=['POST'])
def optimize_route():
    """
    Optimize route using Nearest Neighbor algorithm.
    
    WHY: Core feature - finds efficient route through multiple stops.
    COMPLEXITY: O(n²) from Nearest Neighbor algorithm.
    
    Request JSON:
        - stops: List of {name, latitude, longitude}
    
    Returns:
        JSON: Optimized route with statistics
    
    Example:
        >>> response = client.post('/api/optimize',
        ...     json={'stops': [
        ...         {'name': 'A', 'latitude': 51.5, 'longitude': -0.1},
        ...         {'name': 'B', 'latitude': 51.6, 'longitude': -0.2}
        ...     ]})
        >>> 'optimized_order' in response.json
        True
    """
    data = request.get_json()
    stops = data.get('stops', [])
    
    if len(stops) < 2:
        return jsonify({'error': 'At least 2 stops required'}), 400
    
    # Validate stop format
    for stop in stops:
        if not all(k in stop for k in ['name', 'latitude', 'longitude']):
            return jsonify({'error': 'Invalid stop format'}), 400
    
    # Run optimization
    result = nearest_neighbor_optimization(stops)
    
    # Get AI predictions
    db = get_db()
    model = get_traffic_model(db.cursor())
    predictions = predict_travel_time(
        model, 
        result['total_distance'],
        len(result['optimized_order']) - 2  # Exclude start and end
    )
    
    return jsonify({
        'route': result,
        'predictions': predictions
    })


@app.route('/api/pois', methods=['GET'])
def get_pois():
    """
    Get points of interest near route.
    
    WHY: Helps travelers find amenities along their route.
    COMPLEXITY: O(n) where n is number of stored POIs.
    
    Query params:
        - latitude: Center latitude
        - longitude: Center longitude
        - radius: Search radius in km (default: 5)
        - category: Filter by category (optional)
    
    Returns:
        JSON: List of POIs with details
    
    Example:
        >>> response = client.get('/api/pois?latitude=51.5&longitude=-0.1&radius=2')
        >>> len(response.json) > 0
        True
    """
    lat = float(request.args.get('latitude', 51.5074))
    lon = float(request.args.get('longitude', -0.1278))
    radius = float(request.args.get('radius', 5))
    category = request.args.get('category')
    
    # Sample POIs (in production, would query Overpass API or database)
    sample_pois = [
        {'name': 'Central Hotel', 'category': 'hotel', 'lat': lat + 0.01, 'lon': lon + 0.01, 'rating': 4.5},
        {'name': 'Quick Fuel Station', 'category': 'fuel', 'lat': lat - 0.01, 'lon': lon + 0.02, 'rating': 4.2},
        {'name': 'The Golden Restaurant', 'category': 'restaurant', 'lat': lat + 0.02, 'lon': lon - 0.01, 'rating': 4.7},
        {'name': 'FitLife Gym', 'category': 'gym', 'lat': lat - 0.02, 'lon': lon - 0.02, 'rating': 4.3},
        {'name': 'City Museum', 'category': 'museum', 'lat': lat + 0.03, 'lon': lon, 'rating': 4.8},
        {'name': 'General Hospital', 'category': 'hospital', 'lat': lat - 0.03, 'lon': lon + 0.01, 'rating': 4.6},
        {'name': 'Public Parking', 'category': 'parking', 'lat': lat + 0.01, 'lon': lon - 0.02, 'rating': 3.9},
        {'name': 'Rest Area', 'category': 'rest_area', 'lat': lat - 0.01, 'lon': lon - 0.03, 'rating': 4.0},
    ]
    
    # Filter by distance
    nearby_pois = []
    for poi in sample_pois:
        dist = haversine_distance(lat, lon, poi['lat'], poi['lon'])
        
        if dist <= radius:
            if category is None or poi['category'] == category:
                poi['distance'] = round(dist, 2)
                nearby_pois.append(poi)
    
    # Sort by distance
    nearby_pois.sort(key=lambda x: x['distance'])
    
    return jsonify(nearby_pois)


# ============================================================================
# FLASK ROUTES - EXPORTS
# ============================================================================

@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """
    Export route report as PDF.
    
    WHY: PDF format for professional sharing and printing.
    COMPLEXITY: O(n) where n is report content size.
    
    Request JSON:
        - route: Optimized route data
        - pois: Points of interest
        - predictions: AI predictions
    
    Returns:
        File: PDF download
    
    Example:
        >>> response = client.post('/api/export/pdf', json=report_data)
        >>> response.content_type
        'application/pdf'
    """
    data = request.get_json()
    
    pdf_content = generate_pdf_report(
        data.get('route', {}),
        data.get('pois', []),
        data.get('predictions', {})
    )
    
    return send_file(
        BytesIO(pdf_content),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"route_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """
    Export route report as Excel spreadsheet.
    
    WHY: Excel allows further data manipulation and analysis.
    COMPLEXITY: O(n) where n is number of rows.
    
    Request JSON:
        - route: Route data
        - pois: Points of interest
        - history: Trip history (optional)
    
    Returns:
        File: Excel download
    
    Example:
        >>> response = client.post('/api/export/excel', json=data)
        >>> response.content_type
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    """
    data = request.get_json()
    
    excel_content = generate_excel_report(
        data.get('route', {}),
        data.get('pois', []),
        data.get('history', [])
    )
    
    return send_file(
        BytesIO(excel_content),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"route_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@app.route('/api/export/csv', methods=['POST'])
def export_csv():
    """
    Export route stops as CSV.
    
    WHY: CSV format for easy import into other tools.
    COMPLEXITY: O(n) where n is number of stops.
    
    Request JSON:
        - route: Route with optimized_order
    
    Returns:
        File: CSV download
    
    Example:
        >>> response = client.post('/api/export/csv', json={'route': route_data})
        >>> b'Stop,Name,Latitude,Longitude' in response.data
        True
    """
    data = request.get_json()
    route = data.get('route', {})
    stops = route.get('optimized_order', [])
    
    # Build CSV content
    lines = ['Stop,Name,Latitude,Longitude']
    for i, stop in enumerate(stops):
        lines.append(f"{i+1},{stop['name']},{stop['latitude']},{stop['longitude']}")
    
    csv_content = '\n'.join(lines)
    
    response = make_response(csv_content)
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = \
        f"attachment; filename=route_stops_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return response


@app.route('/api/export/qr', methods=['POST'])
def export_qr():
    """
    Generate QR code for route sharing.
    
    WHY: Quick mobile access and sharing capability.
    COMPLEXITY: O(n) where n is URL length.
    
    Request JSON:
        - url: Route URL to encode
    
    Returns:
        File: PNG image
    
    Example:
        >>> response = client.post('/api/export/qr', 
        ...     json={'url': 'http://example.com/route/123'})
        >>> response.content_type
        'image/png'
    """
    data = request.get_json()
    url = data.get('url', 'https://daily-route-planner.app')
    
    qr_content = generate_qr_code(url)
    
    return send_file(
        BytesIO(qr_content),
        mimetype='image/png',
        as_attachment=True,
        download_name='route_qr_code.png'
    )


# ============================================================================
# FLASK ROUTES - DRIVER MANAGEMENT (Business Only)
# ============================================================================

@app.route('/api/drivers', methods=['GET'])
@role_required(['business'])
def get_drivers():
    """
    Get all drivers managed by business owner.
    
    WHY: Business owners need to track their drivers.
    COMPLEXITY: O(n) where n is number of drivers.
    
    Returns:
        JSON: List of drivers with performance metrics
    
    Example:
        >>> response = client.get('/api/drivers')
        >>> isinstance(response.json, list)
        True
    """
    db = get_db()
    cursor = db.cursor()
    
    business_id = session['user_id']
    
    cursor.execute('''
        SELECT u.id, u.username, u.vehicle_type, 
               dp.total_trips, dp.total_distance, dp.average_rating,
               dp.on_time_percentage
        FROM users u
        LEFT JOIN driver_performance dp ON u.id = dp.driver_id AND dp.business_id = ?
        WHERE u.user_type = 'driver'
    ''', (business_id,))
    
    drivers = [dict(row) for row in cursor.fetchall()]
    
    return jsonify(drivers)


@app.route('/api/drivers/<int:driver_id>/performance', methods=['GET'])
@role_required(['business'])
def get_driver_performance(driver_id):
    """
    Get detailed performance report for specific driver.
    
    WHY: Detailed analytics for performance reviews.
    COMPLEXITY: O(n) where n is number of trips.
    
    Args:
        driver_id: ID of driver to analyze
    
    Returns:
        JSON: Performance metrics and trip history
    
    Example:
        >>> response = client.get('/api/drivers/5/performance')
        >>> 'trips' in response.json
        True
    """
    db = get_db()
    cursor = db.cursor()
    
    business_id = session['user_id']
    
    # Get driver info
    cursor.execute('''
        SELECT id, username, vehicle_type FROM users
        WHERE id = ? AND user_type = 'driver'
    ''', (driver_id,))
    
    driver = cursor.fetchone()
    if not driver:
        return jsonify({'error': 'Driver not found'}), 404
    
    # Get performance summary
    cursor.execute('''
        SELECT total_trips, total_distance, average_rating, on_time_percentage
        FROM driver_performance
        WHERE driver_id = ? AND business_id = ?
    ''', (driver_id, business_id))
    
    perf = cursor.fetchone()
    
    # Get recent trips
    cursor.execute('''
        SELECT distance, duration, num_stops, departure_hour, 
               day_of_week, fuel_consumed, rating, created_at
        FROM trip_reports
        WHERE driver_id = ?
        ORDER BY created_at DESC
        LIMIT 20
    ''', (driver_id,))
    
    trips = [dict(row) for row in cursor.fetchall()]
    
    return jsonify({
        'driver': dict(driver),
        'performance': dict(perf) if perf else {},
        'recent_trips': trips
    })


# ============================================================================
# WEBSOCKET - REAL-TIME GPS TRACKING
# ============================================================================

@sock.route('/ws/driver-location')
def driver_location_websocket(ws):
    """
    WebSocket endpoint for real-time driver location updates.
    
    WHY: WebSocket provides low-latency bidirectional communication
         for live tracking. Alternative: Server-Sent Events (SSE) for
         one-way updates, but WebSocket supports future features.
    COMPLEXITY: O(1) per message - Direct dictionary operations.
    
    Messages:
        - From driver: {"type": "update", "latitude": x, "longitude": y}
        - To tracker: {"driver_id": x, "latitude": y, "longitude": z, "timestamp": t}
    
    Example:
        # Client sends location update
        ws.send(json.dumps({'type': 'update', 'lat': 51.5, 'lon': -0.1}))
    """
    while True:
        try:
            data = ws.receive()
            message = json.loads(data)
            
            if message.get('type') == 'update':
                driver_id = session.get('user_id')
                
                if driver_id:
                    # Store latest location
                    driver_locations[driver_id] = {
                        'latitude': message.get('latitude'),
                        'longitude': message.get('longitude'),
                        'timestamp': datetime.now().isoformat(),
                        'speed': message.get('speed', 0)
                    }
                    
                    # Broadcast to all connected clients (simplified)
                    # In production, use Redis pub/sub for multi-server
                    broadcast_message = {
                        'driver_id': driver_id,
                        'latitude': message.get('latitude'),
                        'longitude': message.get('longitude'),
                        'timestamp': driver_locations[driver_id]['timestamp']
                    }
                    
                    # Send acknowledgment
                    ws.send(json.dumps({
                        'status': 'received',
                        'data': broadcast_message
                    }))
            
        except Exception as e:
            # Log error and continue
            print(f"WebSocket error: {e}")
            break


@app.route('/api/drivers/locations', methods=['GET'])
@role_required(['business'])
def get_all_driver_locations():
    """
    Get current locations of all tracked drivers.
    
    WHY: Business owners need real-time visibility of fleet.
    COMPLEXITY: O(n) where n is number of tracked drivers.
    
    Returns:
        JSON: Dict of driver locations
    
    Example:
        >>> response = client.get('/api/drivers/locations')
        >>> isinstance(response.json, dict)
        True
    """
    return jsonify(driver_locations)


# ============================================================================
# FLASK ROUTES - SAVED LOCATIONS
# ============================================================================

@app.route('/api/locations', methods=['GET', 'POST'])
@login_required
def manage_locations():
    """
    Get or save user's favorite locations.
    
    WHY: Personalization improves user experience.
    COMPLEXITY: O(n) for GET, O(1) for POST.
    
    GET: Returns all saved locations for user
    POST: Saves new location
    
    Returns:
        JSON: List of locations or success confirmation
    
    Example:
        >>> response = client.get('/api/locations')
        >>> isinstance(response.json, list)
        True
    """
    db = get_db()
    cursor = db.cursor()
    user_id = session['user_id']
    
    if request.method == 'GET':
        cursor.execute('''
            SELECT id, name, latitude, longitude, category, created_at
            FROM saved_locations
            WHERE user_id = ?
            ORDER BY created_at DESC
        ''', (user_id,))
        
        locations = [dict(row) for row in cursor.fetchall()]
        return jsonify(locations)
    
    elif request.method == 'POST':
        data = request.get_json()
        
        required = ['name', 'latitude', 'longitude']
        if not all(k in data for k in required):
            return jsonify({'error': 'Missing required fields'}), 400
        
        cursor.execute('''
            INSERT INTO saved_locations (user_id, name, latitude, longitude, category)
            VALUES (?, ?, ?, ?, ?)
        ''', (user_id, data['name'], data['latitude'], 
              data['longitude'], data.get('category')))
        
        db.commit()
        
        return jsonify({'success': True, 'id': cursor.lastrowid})


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == '__main__':
    # Initialize database on startup
    init_db()
    
    # Run Flask development server
    # In production, use Gunicorn or uWSGI
    print("=" * 60)
    print("Daily Route Planner - Starting Server")
    print("=" * 60)
    print("Author: Chioma Okoye")
    print("Year: 2026")
    print("=" * 60)
    print("Demo Accounts:")
    print("  Business Owner: business_owner / pass123")
    print("  Driver: driver_sarah / pass123")
    print("  Traveler: traveler_john / pass123")
    print("=" * 60)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
